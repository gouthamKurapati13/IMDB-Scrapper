from selenium import webdriver
from bs4 import BeautifulSoup as bs
import openpyxl

driver = webdriver.Firefox(executable_path='C:\\Selenium Drivers\\geckodriver.exe')
driver.maximize_window()

excel = openpyxl.Workbook()

def getTop250(url="https://www.imdb.com/chart/top/"):
    driver.get(url)
    source = driver.page_source.encode('utf-8').strip()
    soup = bs(source,'lxml')
    sheet = excel.create_sheet()
    sheet.title = 'Top 250 Movies'
    sheet.append(["Movie Rank","Movie Name","Year of Release","Movie Rating"])
    movies = soup.find('tbody',class_='lister-list').findAll('tr')
    for movie in movies:
        movieSno = movie.find('td',class_='titleColumn').text.strip().split('.')[0]
        movieTitle = movie.find('td',class_='titleColumn').a.text
        movieYear = movie.find('td',class_='titleColumn').span.text.strip('()')
        movieRating = movie.find('td',class_="ratingColumn imdbRating").strong.text
        sheet.append([movieSno,movieTitle,movieYear,movieRating])
    print("Done Top 250...")


def findGenre(url="https://www.imdb.com/chart/top/"):
    driver.get(url)
    source = driver.page_source.encode('utf-8').strip()
    soup = bs(source,'lxml')
    genres = dict()
    gs = soup.find('ul',class_='quicklinks').findAll('li')
    for g in gs:
        genres[g.a.text.strip()] = "https://www.imdb.com"+g.find('a')['href']
    return genres


def getData(url):
    driver.get(url)
    source = driver.page_source.encode('utf-8').strip()
    soup = bs(source,'lxml')
    movies = soup.findAll('div',class_="lister-item mode-advanced")
    for movie in movies:
        movieSno = movie.find('div',class_='lister-item-content').find('span',class_="lister-item-index unbold text-primary").text
        movieTitle = movie.find('div',class_='lister-item-content').a.text
        movieYear = movie.find('div',class_='lister-item-content').find('span',class_="lister-item-year text-muted unbold").text.strip('()')
        movieRating = movie.find('div',class_='lister-item-content').find('div',class_="inline-block ratings-imdb-rating").strong.text
        sheet.append([movieSno,movieTitle,movieYear,movieRating])
    if soup.find('a',class_="lister-page-next next-page"):
            getData('https://www.imdb.com'+soup.find('a',class_="lister-page-next next-page")['href'])

getTop250()
genres = findGenre()

for genre in genres:
    sheet = excel.create_sheet()
    sheet.title = genre
    sheet.append(["Movie Rank","Movie Name","Year of Release","Movie Rating"])
    getData(genres[genre])
    print("Done",genre+'...')


excel.save(r"C:\Users\gouth\OneDrive\Desktop\IMDB Scrapped Data.xlsx")

driver.close()